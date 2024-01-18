import selenium
from tier_values import tier_values_OPGG, average_tier_value
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from openpyxl import Workbook as wb, load_workbook 
import time

def OPGG():
    service = Service()
    option = webdriver.ChromeOptions()
    driver = webdriver.Chrome(service=service, options=option)


    top_tier_list = []
    jg_tier_list = []
    mid_tier_list = []
    adc_tier_list = []
    supp_tier_list = []

    role_tier_lists = {
        'top': top_tier_list,
        'jungle': jg_tier_list, 
        'mid': mid_tier_list, 
        'adc': adc_tier_list, 
        'support': supp_tier_list,
        } 

    url = "https://www.op.gg/champions?position=top"
    driver.get(url)
    time.sleep(4)
    find = driver.find_element(By.XPATH, '//button[@mode="primary"]')
    find.click()
    time.sleep(2)

    average_tier_values = []
    for role, tier_list in role_tier_lists.items():
        url = f"https://www.op.gg/champions?position={role}"
        driver.get(url)
        time.sleep(2)


        champions = []
        tiers = []
        champion_info = []
        elements = driver.find_elements(By.CLASS_NAME, "css-4ee3gn")
        all_tiers = []
        for i in elements:
            champions.append(i.text)
        elements = driver.find_elements(By.CLASS_NAME, "css-1qly9n1")
        for i in elements:
            tiers.append(i.text)
        for i in range(len(champions)):
            champion_info = []
            if champions[i] == 'Nunu & Willump':
                champions[i] = 'Nunu'
            champion_info.append(champions[i])
            champion_info.append(tiers[i])
            champion_info.append(tier_values_OPGG(tiers[i]))
            all_tiers.append(champion_info[2])
            tier_list.append(champion_info)
        average_tier_values.append(average_tier_value(all_tiers))
    with open('LoLChampions.txt') as f:
        contents = f.read().replace('\n', '').split(",")

    top_tier_list.sort(key=lambda x: x[0])
    jg_tier_list.sort(key=lambda x: x[0])
    mid_tier_list.sort(key=lambda x: x[0])
    adc_tier_list.sort(key=lambda x: x[0])
    supp_tier_list.sort(key=lambda x: x[0])  
    
    wb=load_workbook('LoLChampions.xlsx')
    ws=wb['TOP']
    count = 0
    for i in range(0,len(contents)):    
        if contents[i] == (top_tier_list[count][0]).upper():
            ws.cell(row=i+2, column=12).value = top_tier_list[count][1]
            ws.cell(row=i+2, column=13).value = top_tier_list[count][2]/average_tier_values[0]
            if count < len(top_tier_list)-1:
                count += 1
        else:
            ws.cell(row=i+2, column=12).value = "Not Found"
            ws.cell(row=i+2, column=13).value = 0
    count = 0
    ws=wb['JG']
    for i in range(0,len(contents)):
        if contents[i] == (jg_tier_list[count][0]).upper():
            ws.cell(row=i+2, column=12).value = jg_tier_list[count][1]
            ws.cell(row=i+2, column=13).value = jg_tier_list[count][2]/average_tier_values[1]
            if count < len(jg_tier_list)-1:
                count += 1
        else:
            ws.cell(row=i+2, column=12).value = "Not Found"
            ws.cell(row=i+2, column=13).value = 0
    ws=wb['MID']
    count = 0
    for i in range(0,len(contents)): 
        if contents[i] == (mid_tier_list[count][0]).upper():
            ws.cell(row=i+2, column=12).value = mid_tier_list[count][1]
            ws.cell(row=i+2, column=13).value = mid_tier_list[count][2]/average_tier_values[2]
            if count < len(mid_tier_list)-1:
                count += 1
        else:
            ws.cell(row=i+2, column=12).value = "Not Found"
            ws.cell(row=i+2, column=13).value = 0
    ws=wb['ADC']
    count = 0
    for i in range(0,len(contents)):
        if contents[i] == (adc_tier_list[count][0]).upper():
            ws.cell(row=i+2, column=12).value = adc_tier_list[count][1]
            ws.cell(row=i+2, column=13).value = adc_tier_list[count][2]/average_tier_values[3]
            if count < len(adc_tier_list)-1:
                count += 1
        else:
            ws.cell(row=i+2, column=12).value = "Not Found"
            ws.cell(row=i+2, column=13).value = 0
    ws=wb['SUPP']
    count = 0
    for i in range(0,len(contents)):
        if contents[i] == (supp_tier_list[count][0]).upper():
            ws.cell(row=i+2, column=12).value = supp_tier_list[count][1]
            ws.cell(row=i+2, column=13).value = supp_tier_list[count][2]/average_tier_values[4]
            if count < len(supp_tier_list)-1:
                count += 1
        else:
            ws.cell(row=i+2, column=12).value = "Not Found"
            ws.cell(row=i+2, column=13).value = 0
    wb.save('LoLChampions.xlsx')
    wb.close()   