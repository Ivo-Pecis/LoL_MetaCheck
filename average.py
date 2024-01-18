from openpyxl import Workbook, load_workbook 
from openpyxl.utils import get_column_letter

def Average_values():
    wb=load_workbook('LoLChampions.xlsx')
    roles = ['TOP','JG','MID','ADC','SUPP']
    for i in roles:
        ws=wb[i]
        max_row=ws.max_row
        info = []
        for row in range(2,max_row+1):
            for col in range(3,14,2):
                char=get_column_letter(col)
                hour=ws[char + str(row)].value
                info.append(hour)
            ws.cell(row,15).value = sum(info)/len(info)
            info = []    
        wb.save('LoLChampions.xlsx')
        wb.close()
