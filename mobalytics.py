import selenium
from tier_values import tier_values_mobalytics, average_tier_value
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from pynput.keyboard import Key, Controller
from openpyxl import Workbook as wb, load_workbook 
import time
from pynput.keyboard import Key, Controller

service = Service()
option = webdriver.ChromeOptions()
keyboard = Controller()

driver = webdriver.Chrome(service=service, options=option)

url = "https://www.mobalytics.gg"
driver.get(url)
driver.maximize_window()
time.sleep(2)

find = driver.find_element(By.CLASS_NAME, "fc-button-label")
find.click()
time.sleep(1)

find = driver.find_element(By.CLASS_NAME, "gradient-overlay")
find.click()
time.sleep(2)

find = driver.find_element(By.LINK_TEXT, "Tier List")
find.click()
time.sleep(1)

find = driver.find_element(By.LINK_TEXT, "High Elo Tier List")
find.click()
time.sleep(1)

elements = driver.find_elements(By.CLASS_NAME, "m-kvvnsa")
top_tier_list = []
jg_tier_list = []
mid_tier_list = []
adc_tier_list = []
supp_tier_list = []
for i in range (0,15):
    if elements:
        first_element = elements[i]
        tier = first_element.text.split("\n")
        if i/3 < 1:
            for j in range(len(tier)):
                champion = []
                if tier[j] == 'Nunu & Willump':
                    tier[j] = 'Nunu'
                if i%3==0:
                    champion.append(tier[j])
                    champion.append("S")
                    champion.append(tier_values_mobalytics("S"))
                elif i%3==1:
                    champion.append(tier[j])
                    champion.append("A")
                    champion.append(tier_values_mobalytics("A"))
                elif i%3==2:
                    champion.append(tier[j])
                    champion.append("B")
                    champion.append(tier_values_mobalytics("B"))
                top_tier_list.append(champion)
        elif i/3 < 2:
            for j in range(len(tier)):
                champion = []
                if tier[j] == 'Nunu & Willump':
                    tier[j] = 'Nunu'
                if i%3==0:
                    champion.append(tier[j])
                    champion.append("S")
                    champion.append(tier_values_mobalytics("S"))
                elif i%3==1:
                    champion.append(tier[j])
                    champion.append("A")
                    champion.append(tier_values_mobalytics("A"))
                elif i%3==2:
                    champion.append(tier[j])
                    champion.append("B")
                    champion.append(tier_values_mobalytics("B"))
                jg_tier_list.append(champion)
        elif i/3 < 3:
            for j in range(len(tier)):
                champion = []
                if tier[j] == 'Nunu & Willump':
                    tier[j] = 'Nunu'
                if i%3==0:
                    champion.append(tier[j])
                    champion.append("S")
                    champion.append(tier_values_mobalytics("S"))
                elif i%3==1:
                    champion.append(tier[j])
                    champion.append("A")
                    champion.append(tier_values_mobalytics("A"))
                elif i%3==2:
                    champion.append(tier[j])
                    champion.append("B")
                    champion.append(tier_values_mobalytics("B"))
                mid_tier_list.append(champion)
        elif i/3 < 4:
            for j in range(len(tier)):
                champion = []
                if tier[j] == 'Nunu & Willump':
                    tier[j] = 'Nunu'
                if i%3==0:
                    champion.append(tier[j])
                    champion.append("S")
                    champion.append(tier_values_mobalytics("S"))
                elif i%3==1:
                    champion.append(tier[j])
                    champion.append("A")
                    champion.append(tier_values_mobalytics("A"))
                elif i%3==2:
                    champion.append(tier[j])
                    champion.append("B")
                    champion.append(tier_values_mobalytics("B"))
                adc_tier_list.append(champion)
        elif i/3 < 5:
            for j in range(len(tier)):
                champion = []
                if tier[j] == 'Nunu & Willump':
                    tier[j] = 'Nunu'
                if tier[j] == 'Nunu & Willump':
                    tier[j] = 'Nunu'    
                if i%3==0:
                    champion.append(tier[j])
                    champion.append("S")
                    champion.append(tier_values_mobalytics("S"))
                elif i%3==1:
                    champion.append(tier[j])
                    champion.append("A")
                    champion.append(tier_values_mobalytics("A"))
                elif i%3==2:
                    champion.append(tier[j])
                    champion.append("B")
                    champion.append(tier_values_mobalytics("B"))
                supp_tier_list.append(champion)
top_tier_list.sort(key=lambda x: x[0])
jg_tier_list.sort(key=lambda x: x[0])
mid_tier_list.sort(key=lambda x: x[0])
adc_tier_list.sort(key=lambda x: x[0])
supp_tier_list.sort(key=lambda x: x[0])
print(mid_tier_list)

with open('LoLChampions.txt') as f:
    contents = f.read().replace('\n', '').split(",")
wb=load_workbook('LoLChampions.xlsx')
ws=wb['TOP']
count = 0
for i in range(0,len(contents)):    
    if contents[i] == (top_tier_list[count][0]).upper():
        ws.cell(row=i+2, column=4).value = top_tier_list[count][1]
        ws.cell(row=i+2, column=5).value = top_tier_list[count][2]
        if count < len(top_tier_list)-1:
            count += 1
    else:
        ws.cell(row=i+2, column=4).value = "Not Found"
        ws.cell(row=i+2, column=5).value = 0
count = 0
ws=wb['JG']
for i in range(0,len(contents)):
    if contents[i] == (jg_tier_list[count][0]).upper():
        ws.cell(row=i+2, column=4).value = jg_tier_list[count][1]
        ws.cell(row=i+2, column=5).value = jg_tier_list[count][2]
        if count < len(jg_tier_list)-1:
            count += 1
    else:
        ws.cell(row=i+2, column=4).value = "Not Found"
        ws.cell(row=i+2, column=5).value = 0
ws=wb['MID']
count = 0
for i in range(0,len(contents)): 
    if contents[i] == (mid_tier_list[count][0]).upper():
        ws.cell(row=i+2, column=4).value = mid_tier_list[count][1]
        ws.cell(row=i+2, column=5).value = mid_tier_list[count][2]
        if count < len(mid_tier_list)-1:
            count += 1
    else:
        ws.cell(row=i+2, column=4).value = "Not Found"
        ws.cell(row=i+2, column=5).value = 0
ws=wb['ADC']
count = 0
for i in range(0,len(contents)):
    if contents[i] == (adc_tier_list[count][0]).upper():
        ws.cell(row=i+2, column=4).value = adc_tier_list[count][1]
        ws.cell(row=i+2, column=5).value = adc_tier_list[count][2]
        if count < len(adc_tier_list)-1:
            count += 1
    else:
        ws.cell(row=i+2, column=4).value = "Not Found"
        ws.cell(row=i+2, column=5).value = 0
ws=wb['SUPP']
count = 0
for i in range(0,len(contents)):
    if contents[i] == (supp_tier_list[count][0]).upper():
        ws.cell(row=i+2, column=4).value = supp_tier_list[count][1]
        ws.cell(row=i+2, column=5).value = supp_tier_list[count][2]
        if count < len(supp_tier_list)-1:
            count += 1
    else:
        ws.cell(row=i+2, column=4).value = "Not Found"
        ws.cell(row=i+2, column=5).value = 0
wb.save('LoLChampions.xlsx')
wb.close()