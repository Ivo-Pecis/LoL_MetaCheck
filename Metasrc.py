from openpyxl import Workbook, load_workbook 
from tier_values import tier_values_Metasrc, average_tier_value
from selenium import webdriver
from selenium.webdriver.common.by import By
import time

service = webdriver.chrome.service.Service()
option = webdriver.ChromeOptions()
driver = webdriver.Chrome(service=service, options=option)


top_tier_list = []
jg_tier_list = []
mid_tier_list = []
adc_tier_list = []
supp_tier_list = []

role_tier_lists = {
    'top': top_tier_list,
    'jungle': jg_tier_list, 
    'mid': mid_tier_list, 
    'adc': adc_tier_list, 
    'support': supp_tier_list, 
}
top_tier_values = []
jg_tier_values = []
mid_tier_values = []
adc_tier_values = []
supp_tier_values = []

role_tier_values = {
    'top': top_tier_values,
    'jungle': jg_tier_values,
    'mid': mid_tier_values,
    'adc': adc_tier_values,
    'support': supp_tier_values
}

for role, tier_values_list in role_tier_values.items():
    url = f"https://www.metasrc.com/lol/tier-list/{role}"
    driver.get(url)
    time.sleep(0.5)

    # Find all elements with class "_31hylu._hid39z" using Selenium
    elements = driver.find_elements(By.CLASS_NAME, '_31hylu._hid39z')

    for element in elements:
        # Extract the text content of each element
        text_content = element.text

        # Find the index of the opening parenthesis and closing parenthesis
        start_index = text_content.find('(')
        end_index = text_content.find(')', start_index)

        if start_index != -1 and end_index != -1:
            # Extract the substring between the parentheses
            tier_value = text_content[start_index + 1:end_index]

            # Append the tier value to the tier values list
            tier_values_list.append(tier_value)
        else:
            tier_value = "Value not found"


for role, tier_list in role_tier_lists.items():
    url = f"https://www.metasrc.com/lol/tier-list/{role}"
    driver.get(url)
    time.sleep(0.5)
    elements = driver.find_elements(By.CLASS_NAME, "_ate82z")
    
    
    for element in elements:
        tier_name = element.find_element(By.TAG_NAME, "a").text.strip()
        champions = element.find_elements(By.CLASS_NAME, "_95ecnz")
        tier_champions = []

        for champion in champions:
            champion_name = champion.find_element(By.TAG_NAME, "img").get_attribute("alt")
            champion_name = champion_name.split(",")[0].strip()
            
            if role == "top":
                championer = []
                championer.append(champion_name)
                if float(tier_name) > float(top_tier_values[0][:5]):
                    championer.append("S+")
                    championer.append(tier_values_Metasrc(championer[1]))
                elif float(tier_name) > float(top_tier_values[1][:5]):
                    championer.append("S")
                    championer.append(tier_values_Metasrc(championer[1]))
                elif float(tier_name) > float(top_tier_values[2][:5]):
                    championer.append("A")
                    championer.append(tier_values_Metasrc(championer[1]))
                elif float(tier_name) > float(top_tier_values[3][:5]):
                    championer.append("B")
                    championer.append(tier_values_Metasrc(championer[1]))
                elif float(tier_name) > float(top_tier_values[4][:5]):
                    championer.append("C")
                    championer.append(tier_values_Metasrc(championer[1]))
                elif float(tier_name) > float(top_tier_values[5][:5]):
                    championer.append("D")
                    championer.append(tier_values_Metasrc(championer[1]))
                tier_list.append(championer)
            elif role == "jungle":
                championer = []
                championer.append(champion_name)
                if float(tier_name) > float(jg_tier_values[0][:5]):
                    championer.append("S+")
                    championer.append(tier_values_Metasrc(championer[1]))
                elif float(tier_name) > float(jg_tier_values[1][:5]):
                    championer.append("S")
                    championer.append(tier_values_Metasrc(championer[1]))
                elif float(tier_name) > float(jg_tier_values[2][:5]):
                    championer.append("A")
                    championer.append(tier_values_Metasrc(championer[1]))
                elif float(tier_name) > float(jg_tier_values[3][:5]):
                    championer.append("B")
                    championer.append(tier_values_Metasrc(championer[1]))
                elif float(tier_name) > float(jg_tier_values[4][:5]):
                    championer.append("C")
                    championer.append(tier_values_Metasrc(championer[1]))
                elif float(tier_name) > float(jg_tier_values[5][:5]):
                    championer.append("D")
                    championer.append(tier_values_Metasrc(championer[1]))
                tier_list.append(championer)    
            elif role == "mid":
                championer = []
                championer.append(champion_name)
                if float(tier_name) > float(mid_tier_values[0][:5]):
                    championer.append("S+")
                    championer.append(tier_values_Metasrc(championer[1]))
                elif float(tier_name) > float(mid_tier_values[1][:5]):
                    championer.append("S")
                    championer.append(tier_values_Metasrc(championer[1]))
                elif float(tier_name) > float(mid_tier_values[2][:5]):
                    championer.append("A")
                    championer.append(tier_values_Metasrc(championer[1]))
                elif float(tier_name) > float(mid_tier_values[3][:5]):
                    championer.append("B")
                    championer.append(tier_values_Metasrc(championer[1]))
                elif float(tier_name) > float(mid_tier_values[4][:5]):
                    championer.append("C")
                    championer.append(tier_values_Metasrc(championer[1]))
                elif float(tier_name) > float(mid_tier_values[5][:5]):
                    championer.append("D")
                    championer.append(tier_values_Metasrc(championer[1]))
                tier_list.append(championer)    
            elif role == "adc":
                championer = []
                championer.append(champion_name)
                if float(tier_name) > float(adc_tier_values[0][:5]):
                    championer.append("S+")
                    championer.append(tier_values_Metasrc(championer[1]))
                elif float(tier_name) > float(adc_tier_values[1][:5]):
                    championer.append("S")
                    championer.append(tier_values_Metasrc(championer[1]))
                elif float(tier_name) > float(adc_tier_values[2][:5]):
                    championer.append("A")
                    championer.append(tier_values_Metasrc(championer[1]))
                elif float(tier_name) > float(adc_tier_values[3][:5]):
                    championer.append("B")
                    championer.append(tier_values_Metasrc(championer[1]))
                elif float(tier_name) > float(adc_tier_values[4][:5]):
                    championer.append("C")
                    championer.append(tier_values_Metasrc(championer[1]))
                elif float(tier_name) > float(adc_tier_values[5][:5]):
                    championer.append("D")
                    championer.append(tier_values_Metasrc(championer[1]))
                tier_list.append(championer)    
            if role == "support":
                championer = []
                championer.append(champion_name)
                if float(tier_name) > float(supp_tier_values[0][:5]):
                    championer.append("S+")
                    championer.append(tier_values_Metasrc(championer[1]))
                elif float(tier_name) > float(supp_tier_values[1][:5]):
                    championer.append("S")
                    championer.append(tier_values_Metasrc(championer[1]))
                elif float(tier_name) > float(supp_tier_values[2][:5]):
                    championer.append("A")
                    championer.append(tier_values_Metasrc(championer[1]))
                elif float(tier_name) > float(supp_tier_values[3][:5]):
                    championer.append("B")
                    championer.append(tier_values_Metasrc(championer[1]))
                elif float(tier_name) > float(supp_tier_values[4][:5]):
                    championer.append("C")
                    championer.append(tier_values_Metasrc(championer[1]))
                elif float(tier_name) > float(supp_tier_values[5][:5]):
                    championer.append("D")
                    championer.append(tier_values_Metasrc(championer[1]))
                tier_list.append(championer)    

top_tier_list.sort(key=lambda x: x[0])
jg_tier_list.sort(key=lambda x: x[0])
mid_tier_list.sort(key=lambda x: x[0])
adc_tier_list.sort(key=lambda x: x[0])
supp_tier_list.sort(key=lambda x: x[0])
with open('LoLChampions.txt') as f:
    contents = f.read().replace('\n', '').split(",")
wb=load_workbook('LoLChampions.xlsx')
ws=wb['TOP']
count = 0
for i in range(0,len(contents)):
    ws.cell(row=i+2, column=1).value = contents[i]    
    if contents[i] == (top_tier_list[count][0]).upper():
        ws.cell(row=i+2, column=2).value = top_tier_list[count][1]
        ws.cell(row=i+2, column=3).value = top_tier_list[count][2]
        if count < len(top_tier_list)-1:
            count += 1
    else:
        ws.cell(row=i+2, column=2).value = "Not Found"
        ws.cell(row=i+2, column=3).value = 0
count = 0
ws=wb['JG']
for i in range(0,len(contents)):
    ws.cell(row=i+2, column=1).value = contents[i]    
    if contents[i] == (jg_tier_list[count][0]).upper():
        ws.cell(row=i+2, column=2).value = jg_tier_list[count][1]
        ws.cell(row=i+2, column=3).value = jg_tier_list[count][2]
        if count < len(jg_tier_list)-1:
            count += 1
    else:
        ws.cell(row=i+2, column=2).value = "Not Found"
        ws.cell(row=i+2, column=3).value = 0
ws=wb['MID']
count = 0
for i in range(0,len(contents)):
    ws.cell(row=i+2, column=1).value = contents[i]    
    if contents[i] == (mid_tier_list[count][0]).upper():
        ws.cell(row=i+2, column=2).value = mid_tier_list[count][1]
        ws.cell(row=i+2, column=3).value = mid_tier_list[count][2]
        if count < len(mid_tier_list)-1:
            count += 1
    else:
        ws.cell(row=i+2, column=2).value = "Not Found"
        ws.cell(row=i+2, column=3).value = 0
ws=wb['ADC']
count = 0
for i in range(0,len(contents)):
    ws.cell(row=i+2, column=1).value = contents[i]    
    if contents[i] == (adc_tier_list[count][0]).upper():
        ws.cell(row=i+2, column=2).value = adc_tier_list[count][1]
        ws.cell(row=i+2, column=3).value = adc_tier_list[count][2]
        if count < len(adc_tier_list)-1:
            count += 1
    else:
        ws.cell(row=i+2, column=2).value = "Not Found"
        ws.cell(row=i+2, column=3).value = 0
ws=wb['SUPP']
count = 0
for i in range(0,len(contents)):
    ws.cell(row=i+2, column=1).value = contents[i]    
    if contents[i] == (supp_tier_list[count][0]).upper():
        ws.cell(row=i+2, column=2).value = supp_tier_list[count][1]
        ws.cell(row=i+2, column=3).value = supp_tier_list[count][2]
        if count < len(supp_tier_list)-1:
            count += 1
    else:
        ws.cell(row=i+2, column=2).value = "Not Found"
        ws.cell(row=i+2, column=3).value = 0
wb.save('LoLChampions.xlsx')
wb.close()