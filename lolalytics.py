import selenium
from tier_values import tier_values_LoLalytics, average_tier_value
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from pynput.keyboard import Key, Controller
from openpyxl import Workbook as wb, load_workbook 
import time
from pynput.keyboard import Key, Controller

service = Service()
option = webdriver.ChromeOptions()
keyboard = Controller()

driver = webdriver.Chrome(service=service, options=option)

top_tier_list = []
jg_tier_list = []
mid_tier_list = []
adc_tier_list = []
supp_tier_list = []

role_tier_lists = {
    'top': top_tier_list,
    'jungle': jg_tier_list, 
    'middle': mid_tier_list, 
    'bottom': adc_tier_list, 
    'support': supp_tier_list, 
}

url = f"https://lolalytics.com/lol/tierlist/?lane=Top"
driver.get(url)
time.sleep(2)

find = driver.find_element(By.CLASS_NAME, "ncmp__btn")
find.click()
time.sleep(3)

find = driver.find_element(By.CLASS_NAME, "ncmp__btn.ncmp__btn-danger")
find.click()
time.sleep(2)

for role, tier_list in role_tier_lists.items():
    url = f"https://lolalytics.com/lol/tierlist/?lane={role}"
    driver.get(url)
    time.sleep(2)

    elements = driver.find_elements(By.CLASS_NAME, "TierList_list__j33gd")
    if elements:
            for i in range(len(elements)):
                first_element = elements[i]
                tier = first_element.text.split("\n")
                for j in range(len(tier)):
                    tiers = []
                    if tier[j][:1].isalpha():
                        if tier[j] != 'I' and tier[j] != "II" and tier[j] != "III" and tier[j] != "IV":
                            champion_atributes = tier[j].split(" ")
                            if len(champion_atributes) == 2:
                                champion_name = champion_atributes[0]
                                champion_tier = champion_atributes[1]
                            elif len(champion_atributes) == 3:
                                champion_name = champion_atributes[0] + " " + champion_atributes[1]
                                champion_tier = champion_atributes[2]
                            tiers.append(champion_name)
                            tiers.append(champion_tier)
                            tiers.append(tier_values_LoLalytics(champion_tier))
                            tier_list.append(tiers)
top_tier_list.sort(key=lambda x: x[0])
jg_tier_list.sort(key=lambda x: x[0])
mid_tier_list.sort(key=lambda x: x[0])
adc_tier_list.sort(key=lambda x: x[0])
supp_tier_list.sort(key=lambda x: x[0])
print(mid_tier_list)
with open('LoLChampions.txt') as f:
    contents = f.read().replace('\n', '').split(",")
wb=load_workbook('LoLChampions.xlsx')
ws=wb['TOP']
count = 0
for i in range(0,len(contents)):    
    if contents[i] == (top_tier_list[count][0]).upper():
        ws.cell(row=i+2, column=8).value = top_tier_list[count][1]
        ws.cell(row=i+2, column=9).value = top_tier_list[count][2]
        if count < len(top_tier_list)-1:
            count += 1
    else:
        ws.cell(row=i+2, column=8).value = "Not Found"
        ws.cell(row=i+2, column=9).value = 0
count = 0
ws=wb['JG']
for i in range(0,len(contents)):
    if contents[i] == (jg_tier_list[count][0]).upper():
        ws.cell(row=i+2, column=8).value = jg_tier_list[count][1]
        ws.cell(row=i+2, column=9).value = jg_tier_list[count][2]
        if count < len(jg_tier_list)-1:
            count += 1
    else:
        ws.cell(row=i+2, column=8).value = "Not Found"
        ws.cell(row=i+2, column=9).value = 0
ws=wb['MID']
count = 0
for i in range(0,len(contents)): 
    if contents[i] == (mid_tier_list[count][0]).upper():
        ws.cell(row=i+2, column=8).value = mid_tier_list[count][1]
        ws.cell(row=i+2, column=9).value = mid_tier_list[count][2]
        if count < len(mid_tier_list)-1:
            count += 1
    else:
        ws.cell(row=i+2, column=8).value = "Not Found"
        ws.cell(row=i+2, column=9).value = 0
ws=wb['ADC']
count = 0
for i in range(0,len(contents)):
    if contents[i] == (adc_tier_list[count][0]).upper():
        ws.cell(row=i+2, column=8).value = adc_tier_list[count][1]
        ws.cell(row=i+2, column=9).value = adc_tier_list[count][2]
        if count < len(adc_tier_list)-1:
            count += 1
    else:
        ws.cell(row=i+2, column=8).value = "Not Found"
        ws.cell(row=i+2, column=9).value = 0
ws=wb['SUPP']
count = 0
for i in range(0,len(contents)):
    if contents[i] == (supp_tier_list[count][0]).upper():
        ws.cell(row=i+2, column=8).value = supp_tier_list[count][1]
        ws.cell(row=i+2, column=9).value = supp_tier_list[count][2]
        if count < len(supp_tier_list)-1:
            count += 1
    else:
        ws.cell(row=i+2, column=8).value = "Not Found"
        ws.cell(row=i+2, column=9).value = 0
wb.save('LoLChampions.xlsx')
wb.close()
