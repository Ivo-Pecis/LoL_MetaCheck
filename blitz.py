import selenium
from tier_values import tier_values_blitz, average_tier_value
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from openpyxl import Workbook as wb, load_workbook 
import time

def blitz():
    service = Service()
    option = webdriver.ChromeOptions()
    driver = webdriver.Chrome(service=service, options=option)

    url = "https://www.blitz.gg/lol/tierlist"
    driver.get(url)
    time.sleep(8)

    find = driver.find_element(By.XPATH, '//button[@mode="primary"]')
    find.click()
    time.sleep(2)

    top_tier_list = []
    jg_tier_list = []
    mid_tier_list = []
    adc_tier_list = []
    supp_tier_list = []

    role_tier_lists = {
        'Top': top_tier_list,
        'Jungle': jg_tier_list, 
        'Mid': mid_tier_list, 
        'ADC': adc_tier_list, 
        'Support': supp_tier_list,
        } 
    average_tier_values = []
    for role, tier_list in role_tier_lists.items():
        find = driver.find_element(By.XPATH, f'//button[@data-tip="{role}"]')
        find.click()
        time.sleep(2)

        elements = driver.find_elements(By.CLASS_NAME, "⚡97a853a3")
        all_tiers = []
        for i in range (0,4):
            if elements:
                    first_element = elements[i]
                    tier = first_element.text.split("\n")
                    j=0
                    while j < len(tier):
                        champion = []
                        if tier[j] == 'Nunu & Willump':
                            tier[j] = 'Nunu'
                        champion.append(tier[j])
                        if i == 0:
                            champion.append("S")
                            champion.append(tier_values_blitz("S"))
                        elif i == 1:
                            champion.append("A")
                            champion.append(tier_values_blitz("A"))
                        elif i == 2:
                            champion.append("B")
                            champion.append(tier_values_blitz("B"))
                        elif i == 3:
                            champion.append("C")
                            champion.append(tier_values_blitz("C"))
                        tier_list.append(champion)
                        j+=2
                        all_tiers.append(champion[2])
        average_tier_values.append(average_tier_value(all_tiers))

    top_tier_list.sort(key=lambda x: x[0])
    jg_tier_list.sort(key=lambda x: x[0])
    mid_tier_list.sort(key=lambda x: x[0])
    adc_tier_list.sort(key=lambda x: x[0])
    supp_tier_list.sort(key=lambda x: x[0])

    with open('LoLChampions.txt') as f:
        contents = f.read().replace('\n', '').split(",")
    wb=load_workbook('LoLChampions.xlsx')
    ws=wb['TOP']
    count = 0
    for i in range(0,len(contents)):    
        if contents[i] == (top_tier_list[count][0]).upper():
            ws.cell(row=i+2, column=6).value = top_tier_list[count][1]
            ws.cell(row=i+2, column=7).value = top_tier_list[count][2]/average_tier_values[0]
            if count < len(top_tier_list)-1:
                count += 1
        else:
            ws.cell(row=i+2, column=6).value = "Not Found"
            ws.cell(row=i+2, column=7).value = 0
    count = 0
    ws=wb['JG']
    for i in range(0,len(contents)):
        if contents[i] == (jg_tier_list[count][0]).upper():
            ws.cell(row=i+2, column=6).value = jg_tier_list[count][1]
            ws.cell(row=i+2, column=7).value = jg_tier_list[count][2]/average_tier_values[1]
            if count < len(jg_tier_list)-1:
                count += 1
        else:
            ws.cell(row=i+2, column=6).value = "Not Found"
            ws.cell(row=i+2, column=7).value = 0
    ws=wb['MID']
    count = 0
    for i in range(0,len(contents)): 
        if contents[i] == (mid_tier_list[count][0]).upper():
            ws.cell(row=i+2, column=6).value = mid_tier_list[count][1]
            ws.cell(row=i+2, column=7).value = mid_tier_list[count][2]/average_tier_values[2]
            if count < len(mid_tier_list)-1:
                count += 1
        else:
            ws.cell(row=i+2, column=6).value = "Not Found"
            ws.cell(row=i+2, column=7).value = 0
    ws=wb['ADC']
    count = 0
    for i in range(0,len(contents)):
        if contents[i] == (adc_tier_list[count][0]).upper():
            ws.cell(row=i+2, column=6).value = adc_tier_list[count][1]
            ws.cell(row=i+2, column=7).value = adc_tier_list[count][2]/average_tier_values[3]
            if count < len(adc_tier_list)-1:
                count += 1
        else:
            ws.cell(row=i+2, column=6).value = "Not Found"
            ws.cell(row=i+2, column=7).value = 0
    ws=wb['SUPP']
    count = 0
    for i in range(0,len(contents)):
        if contents[i] == (supp_tier_list[count][0]).upper():
            ws.cell(row=i+2, column=6).value = supp_tier_list[count][1]
            ws.cell(row=i+2, column=7).value = supp_tier_list[count][2]/average_tier_values[4]
            if count < len(supp_tier_list)-1:
                count += 1
        else:
            ws.cell(row=i+2, column=6).value = "Not Found"
            ws.cell(row=i+2, column=7).value = 0
    wb.save('LoLChampions.xlsx')
    wb.close()