import selenium
from tier_values import tier_values_UGG, average_tier_value
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from openpyxl import Workbook as wb, load_workbook 
import time

service = Service()
option = webdriver.ChromeOptions()
driver = webdriver.Chrome(service=service, options=option)

top_tier_list = []
jg_tier_list = []
mid_tier_list = []
adc_tier_list = []
supp_tier_list = []

role_tier_lists = {
    'top-lane': top_tier_list,
    'jungle': jg_tier_list, 
    'mid-lane': mid_tier_list, 
    'adc': adc_tier_list, 
    'support': supp_tier_list, 
}
url = f"https://www.u.gg/lol/"
driver.get(url)
time.sleep(2)

find = driver.find_element(By.XPATH, '//button[@mode="primary"]')
find.click()
for role, tier_list in role_tier_lists.items():
    url = f"https://www.u.gg/lol/{role}-tier-list"
    driver.get(url)
    time.sleep(2)


    driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
    elements = driver.find_elements(By.CLASS_NAME, "rt-tr-group")
    if elements:
        for i in range(len(elements)):
            tiers = []
            first_element = elements[i]
            tier = first_element.text.split("\n")
            if tier[1] == 'Nunu & Willump':
                tier[1] = 'Nunu'
            champion_name = tier[1]
            champion_tier = tier[2]
            tiers.append(champion_name)
            tiers.append(champion_tier)
            tiers.append(tier_values_UGG(champion_tier))
            tier_list.append(tiers)
top_tier_list.sort(key=lambda x: x[0])
jg_tier_list.sort(key=lambda x: x[0])
mid_tier_list.sort(key=lambda x: x[0])
adc_tier_list.sort(key=lambda x: x[0])
supp_tier_list.sort(key=lambda x: x[0])
print (top_tier_list)

with open('LoLChampions.txt') as f:
    contents = f.read().replace('\n', '').split(",")
wb=load_workbook('LoLChampions.xlsx')
ws=wb['TOP']
count = 0
for i in range(0,len(contents)):    
    if contents[i] == (top_tier_list[count][0]).upper():
        ws.cell(row=i+2, column=10).value = top_tier_list[count][1]
        ws.cell(row=i+2, column=11).value = top_tier_list[count][2]
        if count < len(top_tier_list)-1:
            count += 1
    else:
        ws.cell(row=i+2, column=10).value = "Not Found"
        ws.cell(row=i+2, column=11).value = 0
count = 0
ws=wb['JG']
for i in range(0,len(contents)):
    if contents[i] == (jg_tier_list[count][0]).upper():
        ws.cell(row=i+2, column=10).value = jg_tier_list[count][1]
        ws.cell(row=i+2, column=11).value = jg_tier_list[count][2]
        if count < len(jg_tier_list)-1:
            count += 1
    else:
        ws.cell(row=i+2, column=10).value = "Not Found"
        ws.cell(row=i+2, column=11).value = 0
ws=wb['MID']
count = 0
for i in range(0,len(contents)): 
    if contents[i] == (mid_tier_list[count][0]).upper():
        ws.cell(row=i+2, column=10).value = mid_tier_list[count][1]
        ws.cell(row=i+2, column=11).value = mid_tier_list[count][2]
        if count < len(mid_tier_list)-1:
            count += 1
    else:
        ws.cell(row=i+2, column=10).value = "Not Found"
        ws.cell(row=i+2, column=11).value = 0
ws=wb['ADC']
count = 0
for i in range(0,len(contents)):
    if contents[i] == (adc_tier_list[count][0]).upper():
        ws.cell(row=i+2, column=10).value = adc_tier_list[count][1]
        ws.cell(row=i+2, column=11).value = adc_tier_list[count][2]
        if count < len(adc_tier_list)-1:
            count += 1
    else:
        ws.cell(row=i+2, column=10).value = "Not Found"
        ws.cell(row=i+2, column=11).value = 0
ws=wb['SUPP']
count = 0
for i in range(0,len(contents)):
    if contents[i] == (supp_tier_list[count][0]).upper():
        ws.cell(row=i+2, column=10).value = supp_tier_list[count][1]
        ws.cell(row=i+2, column=11).value = supp_tier_list[count][2]
        if count < len(supp_tier_list)-1:
            count += 1
    else:
        ws.cell(row=i+2, column=10).value = "Not Found"
        ws.cell(row=i+2, column=11).value = 0
wb.save('LoLChampions.xlsx')
wb.close()   